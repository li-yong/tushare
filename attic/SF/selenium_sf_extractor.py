#!/usr/bin/env python3
"""
Selenium script to extract commands and log files from Salesforce cases.

This script:
1. Reads solved.xlsx file to get SF case links
2. Connects to an existing Chrome browser in debug mode
3. Opens each SF case link
4. Clicks 'Non-Automated Posts' 
5. Reads all posts and extracts commands and log files
6. Updates the Excel file with extracted information

Prerequisites:
- Chrome browser running in debug mode: chrome --remote-debugging-port=9222
- solved.xlsx file exists in out/ directory
"""

# run with on Windows desktop: 
# "C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --profile-directory="Profile 4"

import os
import time
import re
import requests
from typing import List, Dict, Tuple, Optional
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SalesforcePostExtractor:
    """
    Extract commands and log files from Salesforce case posts using Selenium.
    """
    
    def __init__(self, debug_port: int = 9222, timeout: int = 30):
        """
        Initialize the extractor.
        
        Args:
            debug_port: Chrome debug port (default: 9222)
            timeout: Default timeout for web operations in seconds
        """
        self.debug_port = debug_port
        self.timeout = timeout
        self.driver = None
        self.excel_file_path = "./out/solved.xlsx"
        
        # Patterns to identify commands and log files
        self.command_patterns = [
            r'(?:^|\s)([a-zA-Z0-9_\-\.]+(?:\s+[a-zA-Z0-9_\-\.=\'"]+)*)\s*(?:\n|$)',  # General command pattern
            r'sudo\s+([^\n]+)',  # sudo commands
            r'systemctl\s+([^\n]+)',  # systemctl commands
            r'kubectl\s+([^\n]+)',  # kubectl commands
            r'docker\s+([^\n]+)',  # docker commands
            r'cat\s+([^\n]+)',  # cat commands
            r'tail\s+([^\n]+)',  # tail commands
            r'grep\s+([^\n]+)',  # grep commands
            r'ls\s+([^\n]+)',  # ls commands
            r'cd\s+([^\n]+)',  # cd commands
            r'find\s+([^\n]+)',  # find commands
        ]
        
        self.log_file_patterns = [
            r'([/\w\-\.]+\.log)',  # .log files
            r'([/\w\-\.]+\.out)',  # .out files
            r'([/\w\-\.]+\.err)',  # .err files
            r'(/var/log/[^\s]+)',  # /var/log/ paths
            r'(/tmp/[^\s]+\.log)',  # /tmp/ log files
            r'(/home/[^\s]+\.log)',  # home directory log files
            r'(journalctl[^\n]*)',  # journalctl commands/outputs
        ]
    
    def check_chrome_debug_port(self) -> bool:
        """
        Check if Chrome debug port is available and responding.
        
        Returns:
            bool: True if Chrome debug port is accessible, False otherwise
        """
        try:
            response = requests.get(f"http://localhost:{self.debug_port}/json/version", timeout=5)
            if response.status_code == 200:
                version_info = response.json()
                logger.info(f"✅ Chrome debug port {self.debug_port} is accessible")
                logger.info(f"   Browser: {version_info.get('Browser', 'Unknown')}")
                logger.info(f"   WebKit: {version_info.get('WebKit-Version', 'Unknown')}")
                return True
            else:
                logger.error(f"❌ Chrome debug port {self.debug_port} returned status: {response.status_code}")
                return False
        except requests.exceptions.RequestException as e:
            logger.error(f"❌ Cannot connect to Chrome debug port {self.debug_port}: {e}")
            return False
    
    def connect_to_existing_chrome(self) -> bool:
        """
        Connect to existing Chrome browser in debug mode.
        
        Returns:
            bool: True if connection successful, False otherwise
        """
        # First check if Chrome debug port is accessible
        if not self.check_chrome_debug_port():
            logger.error("Chrome debug port is not accessible. Please ensure Chrome is running with:")
            logger.error(f'  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port={self.debug_port} --profile-directory="Profile 4"')
            return False
        
        try:
            chrome_options = Options()
            chrome_options.add_experimental_option("debuggerAddress", f"localhost:{self.debug_port}")
            
            # Windows-friendly options - remove Linux-specific flags
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-background-timer-throttling")
            chrome_options.add_argument("--disable-backgrounding-occluded-windows")
            chrome_options.add_argument("--disable-renderer-backgrounding")
            
            # Try to connect with a longer timeout
            self.driver = webdriver.Chrome(options=chrome_options)
            
            # Test the connection by getting the current page title
            try:
                current_url = self.driver.current_url
                logger.info(f"✅ Connected to existing Chrome browser on port {self.debug_port}")
                logger.info(f"   Current page: {current_url}")
                return True
            except Exception as e:
                logger.error(f"❌ Connected to Chrome but cannot access page: {e}")
                return False
            
        except WebDriverException as e:
            logger.error(f"❌ Failed to connect to Chrome browser: {e}")
            logger.error("Troubleshooting steps:")
            logger.error("1. Make sure Chrome is running with debug mode:")
            logger.error(f'   "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port={self.debug_port} --profile-directory="Profile 4"')
            logger.error("2. Check if the debug port is accessible: http://localhost:9222")
            logger.error("3. Make sure you're logged into Salesforce in the Chrome browser")
            logger.error("4. Try closing and restarting Chrome with debug mode")
            return False
    
    def read_excel_file(self) -> List[Dict]:
        """
        Read the solved.xlsx file and extract case information.
        
        Returns:
            List[Dict]: List of case information with CaseNumber, Subject, and SF_Link
        """
        try:
            if not os.path.exists(self.excel_file_path):
                logger.error(f"❌ Excel file not found: {self.excel_file_path}")
                return []
            
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook.active
            
            cases = []
            for row in range(2, worksheet.max_row + 1):  # Skip header row
                case_number_cell = worksheet.cell(row=row, column=1)
                subject_cell = worksheet.cell(row=row, column=2)
                
                case_number = case_number_cell.value
                subject = subject_cell.value if subject_cell.value else ""
                sf_link = case_number_cell.hyperlink.target if case_number_cell.hyperlink else ""
                
                if case_number and sf_link:
                    cases.append({
                        'row': row,
                        'CaseNumber': case_number,
                        'Subject': subject,
                        'SF_Link': sf_link
                    })
            
            logger.info(f"✅ Read {len(cases)} cases from Excel file")
            return cases
            
        except Exception as e:
            logger.error(f"❌ Failed to read Excel file: {e}")
            return []
    
    def open_sf_case(self, sf_link: str) -> bool:
        """
        Open a Salesforce case link in the browser.
        
        Args:
            sf_link: Salesforce case URL
            
        Returns:
            bool: True if opened successfully, False otherwise
        """
        try:
            logger.info(f"🔗 Opening SF case: {sf_link}")
            self.driver.get(sf_link)
            
            # Wait for page to load
            WebDriverWait(self.driver, self.timeout).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            # Additional wait for Salesforce Lightning components to load
            time.sleep(3)
            return True
            
        except TimeoutException:
            logger.error(f"❌ Timeout loading SF case: {sf_link}")
            return False
        except Exception as e:
            logger.error(f"❌ Failed to open SF case: {e}")
            return False
    
    def click_non_automated_posts(self) -> bool:
        """
        Click on 'Non-Automated Posts' button/link.
        
        Returns:
            bool: True if clicked successfully, False otherwise
        """
        try:
            # Try different selectors for "Non-Automated Posts"
            selectors = [
                "//button[contains(text(), 'Non-Automated Posts')]",
                "//a[contains(text(), 'Non-Automated Posts')]",
                "//span[contains(text(), 'Non-Automated Posts')]",
                "//*[contains(text(), 'Non-Automated Posts')]"
            ]
            
            for selector in selectors:
                try:
                    element = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, selector))
                    )
                    element.click()
                    logger.info("✅ Clicked 'Non-Automated Posts'")
                    time.sleep(2)  # Wait for content to load
                    return True
                except TimeoutException:
                    continue
            
            logger.warning("⚠️  'Non-Automated Posts' button not found, proceeding with all posts")
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to click 'Non-Automated Posts': {e}")
            return True  # Continue even if button not found
    
    def extract_commands_from_text(self, text: str) -> List[str]:
        """
        Extract commands from text using regex patterns.
        
        Args:
            text: Text to search for commands
            
        Returns:
            List[str]: List of extracted commands
        """
        commands = []
        
        logger.info("🔍 DEBUG - Command extraction patterns:")
        for i, pattern in enumerate(self.command_patterns, 1):
            pattern_matches = []
            matches = re.finditer(pattern, text, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                command = match.group(1).strip()
                logger.info(f"      🔎 Pattern {i} raw match: '{match.group(0)}' -> extracted: '{command}'")
                if command and len(command) > 2:  # Filter out very short matches
                    commands.append(command)
                    pattern_matches.append(command)
                else:
                    logger.info(f"      🚫 Filtered out (too short or empty): '{command}'")
            
            logger.info(f"   Pattern {i}: {pattern}")
            if pattern_matches:
                logger.info(f"      ✅ Found {len(pattern_matches)} matches: {pattern_matches}")
            else:
                logger.info(f"      ❌ No matches")
        
        # Remove duplicates while preserving order
        unique_commands = []
        for cmd in commands:
            if cmd not in unique_commands:
                unique_commands.append(cmd)
        
        logger.info(f"🔍 DEBUG - Total unique commands after deduplication: {len(unique_commands)}")
        
        return unique_commands
    
    def extract_log_files_from_text(self, text: str) -> List[str]:
        """
        Extract log file references from text using regex patterns.
        
        Args:
            text: Text to search for log files
            
        Returns:
            List[str]: List of extracted log file paths
        """
        log_files = []
        
        logger.info("🔍 DEBUG - Log file extraction patterns:")
        for i, pattern in enumerate(self.log_file_patterns, 1):
            pattern_matches = []
            matches = re.finditer(pattern, text, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                log_file = match.group(1).strip()
                if log_file:
                    log_files.append(log_file)
                    pattern_matches.append(log_file)
            
            logger.info(f"   Pattern {i}: {pattern}")
            if pattern_matches:
                logger.info(f"      ✅ Found {len(pattern_matches)} matches: {pattern_matches}")
            else:
                logger.info(f"      ❌ No matches")
        
        # Remove duplicates while preserving order
        unique_log_files = []
        for log in log_files:
            if log not in unique_log_files:
                unique_log_files.append(log)
        
        logger.info(f"🔍 DEBUG - Total unique log files after deduplication: {len(unique_log_files)}")
        
        return unique_log_files
    
    def read_all_posts(self) -> Tuple[List[str], List[str]]:
        """
        Read all posts from the current Salesforce case page.
        
        Returns:
            Tuple[List[str], List[str]]: (commands, log_files)
        """
        try:
            # Wait for posts to load
            time.sleep(3)
            
            # First try to find the specific chatter feeds container with flexible selectors
            container_element = None
            
            # Try flexible selectors (most specific to least specific)
            container_selectors = [
                "c-vast-chatter-feeds-filtered-by-tab",  # Most specific: target the component directly
                "c-vast-chatter > div > c-vast-chatter-feeds-filtered-by-tab",  # Within c-vast-chatter
                "[id^='tab-'] c-vast-chatter-feeds-filtered-by-tab",  # Any tab with the component
                "slot > flexipage-component2 > slot > c-vast-chatter > div > c-vast-chatter-feeds-filtered-by-tab"  # Without specific tab ID
            ]
            
            for selector in container_selectors:
                try:
                    container_element = self.driver.find_element(By.CSS_SELECTOR, selector)
                    logger.info(f"✅ Found specific chatter feeds container using: {selector}")
                    break
                except NoSuchElementException:
                    continue
            
            if not container_element:
                logger.warning("⚠️  Specific chatter feeds container not found with any selector, searching entire page")
            
            # Try different selectors for posts/comments
            post_selectors = [
                ".forceChatterFeedItemPikeBody",
                ".feeditemtext",
                ".slds-rich-text-editor__output",
                ".uiOutputRichText",
                "[data-aura-class='forceChatterFeedItemPikeBody']",
                ".cuf-body",
                ".feedBodyRender"
            ]
            
            all_text = ""
            posts_found = False
            
            # Search within the specific container if found, otherwise search the entire page
            search_context = container_element if container_element else self.driver
            
            for selector in post_selectors:
                try:
                    posts = search_context.find_elements(By.CSS_SELECTOR, selector)
                    if posts:
                        scope_msg = "within specific container" if container_element else "on entire page"
                        logger.info(f"✅ Found {len(posts)} posts using selector: {selector} ({scope_msg})")
                        for post in posts:
                            post_text = post.text.strip()
                            if post_text:
                                all_text += post_text + "\n\n"
                        posts_found = True
                        break
                except Exception as e:
                    continue
            
            if not posts_found:
                # Fallback: get all text from container or page
                if container_element:
                    logger.warning("⚠️  No posts found with standard selectors in container, using container text")
                    all_text = container_element.text
                else:
                    logger.warning("⚠️  No posts found with standard selectors, using page text")
                    all_text = self.driver.find_element(By.TAG_NAME, "body").text
            
            # Debug: Log the raw extracted text
            logger.info("🔍 DEBUG - Raw text extracted from page:")
            logger.info("-" * 80)
            logger.info(all_text[:2000] + ("..." if len(all_text) > 2000 else ""))  # First 2000 chars
            logger.info("-" * 80)
            logger.info(f"📏 Total text length: {len(all_text)} characters")
            
            # Extract commands and log files
            commands = self.extract_commands_from_text(all_text)
            log_files = self.extract_log_files_from_text(all_text)
            
            # Debug: Log what was sent to extraction methods
            logger.info("🔍 DEBUG - Text sent to command extraction:")
            logger.info("-" * 50)
            logger.info(all_text[:1000] + ("..." if len(all_text) > 1000 else ""))  # First 1000 chars
            logger.info("-" * 50)
            
            scope_info = "from specific container" if container_element else "from entire page"
            logger.info(f"✅ Extracted {len(commands)} commands and {len(log_files)} log files {scope_info}")
            
            # Log the actual extracted commands
            if commands:
                logger.info("📋 Extracted Commands:")
                for i, cmd in enumerate(commands, 1):
                    logger.info(f"   {i}. {cmd}")
            else:
                logger.info("📋 No commands extracted")
            
            # Log the actual extracted log files
            if log_files:
                logger.info("📄 Extracted Log Files:")
                for i, log_file in enumerate(log_files, 1):
                    logger.info(f"   {i}. {log_file}")
            else:
                logger.info("📄 No log files extracted")
            
            return commands, log_files
            
        except Exception as e:
            logger.error(f"❌ Failed to read posts: {e}")
            return [], []
    
    def update_excel_file(self, case_data: List[Dict]) -> bool:
        """
        Update the Excel file with extracted commands and log files.
        
        Args:
            case_data: List of case data with extracted information
            
        Returns:
            bool: True if updated successfully, False otherwise
        """
        try:
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook.active
            
            # Add new headers if they don't exist
            if worksheet.max_column < 3:
                worksheet.cell(row=1, column=3, value="Commands").font = Font(bold=True)
            if worksheet.max_column < 4:
                worksheet.cell(row=1, column=4, value="Log Files").font = Font(bold=True)
            
            # Update each case row
            for case in case_data:
                row = case['row']
                commands = "; ".join(case.get('commands', []))
                log_files = "; ".join(case.get('log_files', []))
                
                # Update commands column
                commands_cell = worksheet.cell(row=row, column=3, value=commands)
                
                # Update log files column
                log_files_cell = worksheet.cell(row=row, column=4, value=log_files)
            
            # Auto-adjust column widths
            for col in range(1, 5):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in worksheet[column_letter]:
                    try:
                        if row.value and len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                
                # Set column width with some padding
                adjusted_width = min(max_length + 2, 80)  # Cap at 80 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(self.excel_file_path)
            logger.info(f"✅ Updated Excel file: {self.excel_file_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to update Excel file: {e}")
            return False
    
    def process_all_cases(self) -> bool:
        """
        Process all cases from the Excel file.
        
        Returns:
            bool: True if processing completed successfully, False otherwise
        """
        try:
            # Connect to Chrome browser
            if not self.connect_to_existing_chrome():
                return False
            
            # Read Excel file
            cases = self.read_excel_file()
            if not cases:
                return False
            
            # Process each case
            processed_cases = []
            for i, case in enumerate(cases, 1):
                logger.info(f"\n📋 Processing case {i}/{len(cases)}: {case['CaseNumber']}")
                
                # Open SF case
                if not self.open_sf_case(case['SF_Link']):
                    case['commands'] = []
                    case['log_files'] = []
                    processed_cases.append(case)
                    continue
                
                # Click Non-Automated Posts
                self.click_non_automated_posts()
                
                # Read all posts and extract information
                commands, log_files = self.read_all_posts()
                
                case['commands'] = commands
                case['log_files'] = log_files
                processed_cases.append(case)
                
                logger.info(f"   Commands: {len(commands)}")
                logger.info(f"   Log files: {len(log_files)}")
                
                # Add delay between cases to avoid overwhelming the server
                time.sleep(2)
            
            # Update Excel file with all extracted data
            self.update_excel_file(processed_cases)
            
            # Summary
            total_commands = sum(len(case.get('commands', [])) for case in processed_cases)
            total_log_files = sum(len(case.get('log_files', [])) for case in processed_cases)
            
            logger.info(f"\n📊 Processing Summary:")
            logger.info(f"   Cases processed: {len(processed_cases)}")
            logger.info(f"   Total commands extracted: {total_commands}")
            logger.info(f"   Total log files extracted: {total_log_files}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to process cases: {e}")
            return False
        
        finally:
            if self.driver:
                # Don't close the browser, just disconnect
                self.driver.quit()
                logger.info("🔌 Disconnected from Chrome browser")

def main():
    """
    Main function to run the Salesforce post extractor.
    """
    print("🚀 Salesforce Post Extractor")
    print("=" * 50)
    print("This script will:")
    print("1. Connect to existing Chrome browser (debug mode)")
    print("2. Read solved.xlsx file")
    print("3. Open each SF case link")
    print("4. Extract commands and log files from posts")
    print("5. Update Excel file with extracted data")
    print()
    
    # Check if Excel file exists
    excel_path = "./out/solved.xlsx"
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        print("Please make sure the solved.xlsx file exists in the out/ directory.")
        return
    
    # Check Chrome debug instructions
    print("📋 Prerequisites:")
    print("   - Chrome browser must be running in debug mode:")
    print("     Windows: \"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe\" --remote-debugging-port=9222 --profile-directory=\"Profile 4\"")
    print("     Linux/Mac: google-chrome --remote-debugging-port=9222")
    print("   - Make sure you're logged into Salesforce in the browser")
    print("   - Verify debug mode: Open http://localhost:9222 in another browser tab")
    print()
    

    
    input("Press Enter when ready to continue...")
    
    # Run extractor
    extractor = SalesforcePostExtractor()
    success = extractor.process_all_cases()
    
    if success:
        print("\n✅ Processing completed successfully!")
        print(f"📄 Updated file: {excel_path}")
    else:
        print("\n❌ Processing failed. Please check the logs above.")

if __name__ == "__main__":
    main()
