import requests
import json
import csv
import os
import argparse
from datetime import datetime
from typing import List, Dict, Optional
from dotenv import load_dotenv
import openai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class SalesforceCasesClient:
    """
    A client to retrieve Salesforce Cases via REST API and save them to files.
    """
    
    def __init__(self, username: str, password: str, security_token: str, 
                 domain: str = "login", is_sandbox: bool = False, enable_ai_cleaning: bool = False):
        """
        Initialize the Salesforce client.
        
        Args:
            username: Your Salesforce username
            password: Your Salesforce password
            security_token: Your Salesforce security token
            domain: 'login' for production, 'test' for sandbox, or custom domain
            is_sandbox: True if using sandbox environment
            enable_ai_cleaning: Enable AI-powered subject cleaning
        """
        self.username = username
        self.password = password
        self.security_token = security_token
        self.enable_ai_cleaning = enable_ai_cleaning
        
        if is_sandbox:
            self.base_url = "https://test.salesforce.com"
        else:
            # Use your custom domain for authentication
            self.base_url = "https://vastdata.my.salesforce.com"
            
        self.session_id = None
        self.instance_url = None
        self.config_file = "salesforce_config.json"
        
        # Initialize OpenAI client if AI cleaning is enabled
        self.openai_client = None
        if self.enable_ai_cleaning:
            self._init_openai_client()
    
    def _init_openai_client(self) -> bool:
        """Initialize OpenAI client with API key from environment."""
        try:
            # Load environment variables
            load_dotenv('.env/openai.env')
            api_key = os.getenv('OPENAI_API_KEY')
            
            if not api_key:
                print("⚠️  Warning: OPENAI_API_KEY not found in environment. AI cleaning disabled.")
                self.enable_ai_cleaning = False
                return False
            
            self.openai_client = openai.OpenAI(api_key=api_key)
            print("✅ OpenAI client initialized for AI subject cleaning")
            return True
            
        except Exception as e:
            print(f"⚠️  Warning: Failed to initialize OpenAI client: {e}. AI cleaning disabled.")
            self.enable_ai_cleaning = False
            return False
    
    def clean_subject_with_ai(self, subject: str) -> str:
        """
        Clean a case subject using OpenAI to create a brief, clean title.
        
        Args:
            subject: Original case subject
            
        Returns:
            str: Cleaned subject or original if AI cleaning fails
        """
        if not self.enable_ai_cleaning or not self.openai_client or not subject:
            return subject
        
        try:
            prompt = f"""For each line of input (representing a Salesforce incident report subject), generate a concise title describing only the nature of the issue.

Ignore customer names.

Remove all serial numbers, IP addresses, timestamps, case numbers, and other identifiers.

Summarize the problem type in very brief, plain text.

Output one clean title per line.

The line to be proceeded is: {subject}"""
            
            response = self.openai_client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "user", "content": prompt}
                ],
                max_tokens=100,
                temperature=0.3
            )
            
            cleaned_subject = response.choices[0].message.content.strip()
            
            # Fallback to original if cleaning result is empty or too short
            if len(cleaned_subject) < 5:
                return subject
                
            return cleaned_subject
            
        except Exception as e:
            print(f"⚠️  Warning: AI subject cleaning failed for '{subject[:50]}...': {e}")
            return subject
    
    def clean_subjects_batch(self, subjects: List[str]) -> List[str]:
        """
        Clean multiple subjects in a batch for efficiency.
        
        Args:
            subjects: List of original subjects
            
        Returns:
            List[str]: List of cleaned subjects
        """
        if not self.enable_ai_cleaning or not self.openai_client or not subjects:
            return subjects
        
        try:
            # Create batch prompt
            subjects_text = "\n".join([f"{i+1}. {subject}" for i, subject in enumerate(subjects)])
            
            prompt = f"""For each line of input (representing a Salesforce incident report subject), generate a concise title describing only the nature of the issue.

Ignore customer names.

Remove all serial numbers, IP addresses, timestamps, case numbers, and other identifiers.

Summarize the problem type in very brief, plain text.

Output one clean title per line, maintaining the same order.

The lines to be processed are:
{subjects_text}"""
            
            response = self.openai_client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "user", "content": prompt}
                ],
                max_tokens=500,
                temperature=0.3
            )
            
            cleaned_text = response.choices[0].message.content.strip()
            cleaned_subjects = []
            
            # Parse the response - expect numbered or plain lines
            lines = cleaned_text.split('\n')
            for i, line in enumerate(lines):
                # Remove numbering if present
                clean_line = line.strip()
                if clean_line.startswith(f"{i+1}."):
                    clean_line = clean_line[len(f"{i+1}."):].strip()
                
                # Fallback to original if cleaning result is empty or too short
                if len(clean_line) < 5 and i < len(subjects):
                    clean_line = subjects[i]
                
                cleaned_subjects.append(clean_line)
            
            # Ensure we have the same number of results
            while len(cleaned_subjects) < len(subjects):
                cleaned_subjects.append(subjects[len(cleaned_subjects)])
            
            print(f"✅ AI cleaned {len(subjects)} case subjects")
            return cleaned_subjects[:len(subjects)]
            
        except Exception as e:
            print(f"⚠️  Warning: Batch AI subject cleaning failed: {e}")
            return subjects
        
    def load_config(self) -> Dict:
        """Load configuration from JSON file."""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"⚠️  Warning: Could not load config file: {e}")
        return {}
    
    def save_config(self, config: Dict) -> bool:
        """Save configuration to JSON file."""
        try:
            with open(self.config_file, 'w') as f:
                json.dump(config, f, indent=2)
            print(f"✅ Configuration saved to {self.config_file}")
            return True
        except Exception as e:
            print(f"❌ Failed to save config: {e}")
            return False
    
    def clear_config(self) -> bool:
        """Clear saved configuration file."""
        try:
            if os.path.exists(self.config_file):
                os.remove(self.config_file)
                print(f"✅ Configuration file {self.config_file} removed")
            return True
        except Exception as e:
            print(f"❌ Failed to clear config: {e}")
            return False
        
    def authenticate(self) -> bool:
        """
        Authenticate with Salesforce using session ID approach.
        Since username-password flow is not supported, we'll use a different method.
        
        Returns:
            bool: True if authentication successful, False otherwise
        """
        # Load existing configuration
        config = self.load_config()
        
        # Check if we have a saved session ID
        if config.get('session_id') and config.get('instance_url'):
            print("🔍 Found saved session ID, testing...")
            self.session_id = config['session_id']
            self.instance_url = config['instance_url']
            
            # Test the session ID by making a simple query
            if self._test_session():
                print("✅ Using saved session ID")
                return True
            else:
                print("❌ Saved session ID is invalid or expired")
        
        # Need to get a new session ID
        print("📝 Username-password OAuth flow is not supported by your Salesforce org.")
        print("")
        print("🔧 To get a session ID from your browser:")
        print("   1. Log into Salesforce in your browser")
        print("   2. Open Developer Tools (F12)")
        print("   3. Go to Application/Storage tab")
        print("   4. Look for 'sid' cookie value")
        print("   5. Copy the session ID value")
        print("")
        
        session_id = input("Enter your Salesforce session ID (or press Enter to skip): ").strip()
        
        if session_id:
            self.session_id = session_id
            self.instance_url = self.base_url
            
            # Test the new session ID
            if self._test_session():
                print(f"✅ Session ID is valid")
                print(f"   Instance URL: {self.instance_url}")
                
                # Save the configuration
                config = {
                    'session_id': self.session_id,
                    'instance_url': self.instance_url,
                    'last_updated': datetime.now().isoformat()
                }
                self.save_config(config)
                return True
            else:
                print("❌ Invalid session ID provided")
                return False
        else:
            print("❌ No session ID provided. Cannot proceed.")
            return False
    
    def _test_session(self) -> bool:
        """Test if the session ID is valid by making a simple API call."""
        if not self.session_id or not self.instance_url:
            return False
        
        headers = {
            'Authorization': f'Bearer {self.session_id}',
            'Content-Type': 'application/json'
        }
        
        # Simple query to test session validity
        url = f"{self.instance_url}/services/data/v58.0/query"
        params = {'q': 'SELECT Id FROM Case LIMIT 1'}
        
        try:
            response = requests.get(url, headers=headers, params=params)
            return response.status_code == 200
        except:
            return False
    
    def get_cases(self, limit: Optional[int] = None, 
                  status_filter: Optional[str] = None,
                  exclude_merged: bool = False,
                  fields: Optional[List[str]] = None,
                  include_comments: bool = True,
                  include_history: bool = False) -> List[Dict]:
        """
        Retrieve cases from Salesforce.
        
        Args:
            limit: Maximum number of cases to retrieve
            status_filter: Filter by case status (e.g., 'New', 'Working', 'Escalated')
            fields: List of fields to retrieve. If None, gets common fields.
            include_comments: Include case comments (CaseComments)
            include_history: Include case history (CaseHistory) - field changes
            
        Returns:
            List of case dictionaries with comments and history if requested
        """
        if not self.session_id:
            print("❌ Not authenticated. Call authenticate() first.")
            return []
        
        # Default fields if none specified - include key custom fields
        if fields is None:
            fields = [
                'Id', 'CaseNumber', 'Subject', 'Description', 'Status', 
                'Priority', 'Origin', 'Type', 'CreatedDate', 'LastModifiedDate',
                'ClosedDate', 'AccountId', 'ContactId', 'OwnerId', 'Reason',
                'SuppliedName', 'SuppliedEmail',
                # Key custom description fields
                'Detailed_Description__c', 'Brief__c', 'Customer_Notes__c',
                'Problem_Description__c', 'Notes__c', 'Handover_notes__c'
            ]
        
        # Build SOQL query
        fields_str = ', '.join(fields)
        query = f"SELECT {fields_str} FROM Case"
        
        conditions = []
        if status_filter:
            conditions.append(f"Status = '{status_filter}'")
        
        if exclude_merged:
            conditions.append("(NOT Subject LIKE '%Merged%')")
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY CreatedDate DESC"
        
        if limit:
            query += f" LIMIT {limit}"
        
        # Get cases first
        cases = self._execute_query(query)
        
        if cases:
            case_ids = [case['Id'] for case in cases]
            
            # Get complete Chatter feed data
            print("🔍 Retrieving Chatter feed...")
            feed_items = self._get_case_feed(case_ids)
            
            print("🔍 Retrieving feed comments...")
            feed_comments = self._get_feed_comments(case_ids)
            
            print("🔍 Retrieving feed tracked changes...")
            feed_tracked_changes = self._get_feed_tracked_changes(case_ids)
            
            print("🔍 Retrieving content documents...")
            content_documents = self._get_case_content_documents(case_ids)
            
            # Get all case-related objects
            print("🔍 Retrieving case articles...")
            articles = self._get_case_articles(case_ids)
            
            print("🔍 Retrieving case contact roles...")
            contact_roles = self._get_case_contact_roles(case_ids)
            
            print("🔍 Retrieving case milestones...")
            milestones = self._get_case_milestones(case_ids)
            
            print("🔍 Retrieving case participants...")
            participants = self._get_case_participants(case_ids)
            
            print("🔍 Retrieving case shares...")
            shares = self._get_case_shares(case_ids)
            
            print("🔍 Retrieving case solutions...")
            solutions = self._get_case_solutions(case_ids)
            
            print("🔍 Retrieving case tags...")
            tags = self._get_case_tags(case_ids)
            
            if include_comments:
                print("🔍 Retrieving case comments...")
                comments = self._get_case_comments(case_ids)
                
            if include_history:
                print("🔍 Retrieving case history...")
                history = self._get_case_history(case_ids)
            
            # Attach all data to cases
            for case in cases:
                case_id = case['Id']
                
                # Get feed items for this case and enrich with comments and tracked changes
                case_feed_items = feed_items.get(case_id, [])
                
                # Enrich each feed item with its comments and tracked changes
                for feed_item in case_feed_items:
                    feed_item_id = feed_item['Id']
                    feed_item['Comments'] = feed_comments.get(feed_item_id, [])
                    feed_item['TrackedChanges'] = feed_tracked_changes.get(feed_item_id, [])
                
                case['ChatterFeed'] = case_feed_items
                case['FeedItemCount'] = len(case_feed_items)
                case['ContentDocuments'] = content_documents.get(case_id, [])
                
                # Add all case-related objects
                case['Articles'] = articles.get(case_id, [])
                case['ContactRoles'] = contact_roles.get(case_id, [])
                case['Milestones'] = milestones.get(case_id, [])
                case['Participants'] = participants.get(case_id, [])
                case['Shares'] = shares.get(case_id, [])
                case['Solutions'] = solutions.get(case_id, [])
                case['Tags'] = tags.get(case_id, [])
                
                if include_comments:
                    case['Comments'] = comments.get(case_id, [])
                
                if include_history:
                    case['History'] = history.get(case_id, [])
        
        # Apply AI cleaning to subjects if enabled
        if self.enable_ai_cleaning and cases:
            print("🤖 Applying AI cleaning to case subjects...")
            subjects = [case.get('Subject', '') for case in cases]
            cleaned_subjects = self.clean_subjects_batch(subjects)
            
            for i, case in enumerate(cases):
                if i < len(cleaned_subjects):
                    case['OriginalSubject'] = case.get('Subject', '')
                    case['Subject'] = cleaned_subjects[i]
        
        print(f"✅ Retrieved {len(cases)} cases with additional data")
        return cases
    
    def describe_case_object(self):
        """Describe the Case object to find all available fields."""
        if not self.session_id:
            print("❌ Not authenticated. Call authenticate() first.")
            return
        
        headers = {
            'Authorization': f'Bearer {self.session_id}',
            'Content-Type': 'application/json'
        }
        
        url = f"{self.instance_url}/services/data/v58.0/sobjects/Case/describe"
        
        try:
            print("🔍 Describing Case object to find custom fields...")
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            describe_result = response.json()
            fields = describe_result.get('fields', [])
            
            # Look for fields that might contain description information
            description_fields = []
            for field in fields:
                field_name = field.get('name', '')
                field_label = field.get('label', '')
                field_type = field.get('type', '')
                
                # Look for fields with description-related names
                if any(keyword in field_name.lower() for keyword in ['brief', 'description', 'detail', 'summary', 'note']):
                    description_fields.append({
                        'name': field_name,
                        'label': field_label,
                        'type': field_type,
                        'custom': field.get('custom', False)
                    })
            
            print(f"📋 Found {len(description_fields)} potential description fields:")
            for field in description_fields:
                custom_indicator = " (CUSTOM)" if field['custom'] else ""
                print(f"   {field['name']} - '{field['label']}' ({field['type']}){custom_indicator}")
            
            return description_fields
            
        except requests.exceptions.RequestException as e:
            print(f"❌ Failed to describe Case object: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response: {e.response.text}")
            return []
    
    def get_cases_list_only(self, limit: Optional[int] = None, 
                           status_filter: Optional[str] = None,
                           exclude_merged: bool = False) -> List[Dict]:
        """
        Get basic case list only (no detailed data, feed, or custom fields).
        
        Args:
            limit: Maximum number of cases to retrieve
            status_filter: Filter by case status
            exclude_merged: Exclude cases with "Merged" in subject
            
        Returns:
            List of cases with basic information only
        """
        if not self.session_id:
            print("❌ Not authenticated. Call authenticate() first.")
            return []
        
        # Basic fields only for case list
        fields = [
            'Id', 'CaseNumber', 'Subject', 'Status', 'Priority', 
            'Origin', 'Type', 'CreatedDate', 'LastModifiedDate',
            'ClosedDate', 'AccountId', 'ContactId', 'OwnerId'
        ]
        
        # Build SOQL query
        fields_str = ', '.join(fields)
        query = f"SELECT {fields_str} FROM Case"
        
        conditions = []
        if status_filter:
            conditions.append(f"Status = '{status_filter}'")
        
        if exclude_merged:
            conditions.append("(NOT Subject LIKE '%Merged%')")
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY CreatedDate DESC"
        
        if limit:
            query += f" LIMIT {limit}"
        
        # Get cases only (no additional data)
        cases = self._execute_query(query)
        
        # Apply AI cleaning to subjects if enabled
        if self.enable_ai_cleaning and cases:
            print("🤖 Applying AI cleaning to case subjects...")
            subjects = [case.get('Subject', '') for case in cases]
            cleaned_subjects = self.clean_subjects_batch(subjects)
            
            for i, case in enumerate(cases):
                if i < len(cleaned_subjects):
                    case['OriginalSubject'] = case.get('Subject', '')
                    case['Subject'] = cleaned_subjects[i]
        
        print(f"✅ Retrieved {len(cases)} cases (list only)")
        return cases
    
    def get_cases_by_number(self, case_number: str, 
                           include_comments: bool = True,
                           include_history: bool = False,
                           custom_description_fields: List[Dict] = None) -> List[Dict]:
        """
        Retrieve a specific case by case number.
        
        Args:
            case_number: The case number (e.g., '00078683')
            include_comments: Include case comments (CaseComments)
            include_history: Include case history (CaseHistory) - field changes
            
        Returns:
            List containing the single case with all related data
        """
        if not self.session_id:
            print("❌ Not authenticated. Call authenticate() first.")
            return []
        
        # Base standard fields
        base_fields = [
            'Id', 'CaseNumber', 'Subject', 'Description', 'Status', 
            'Priority', 'Origin', 'Type', 'CreatedDate', 'LastModifiedDate',
            'ClosedDate', 'AccountId', 'ContactId', 'OwnerId', 'Reason',
            'SuppliedName', 'SuppliedEmail'
        ]
        
        # Start with base fields
        fields = base_fields.copy()
        
        # Add discovered custom description fields
        if custom_description_fields:
            for field in custom_description_fields:
                field_name = field.get('name')
                if field_name and field_name not in fields:
                    fields.append(field_name)
                    print(f"📋 Adding custom field: {field_name} - '{field.get('label', '')}'")
        else:
            print("⚠️  No custom description fields discovered")
        
        # Build SOQL query for specific case number
        fields_str = ', '.join(fields)
        query = f"""
        SELECT {fields_str} 
        FROM Case 
        WHERE CaseNumber = '{case_number}'
        """
        
        print(f"🔍 Searching for case number: {case_number}")
        cases = self._execute_query(query)
        
        if not cases:
            print(f"❌ No case found with number: {case_number}")
            return []
        
        # Check what fields were actually returned
        case = cases[0]
        description_info = []
        
        # Standard fields
        if 'Subject' in case and case['Subject']:
            description_info.append(f"Subject: {case['Subject'][:100]}...")
        if 'Description' in case and case['Description']:
            description_info.append(f"Description: {str(case['Description'])[:100]}...")
        
        # Custom description fields
        if custom_description_fields:
            for field in custom_description_fields:
                field_name = field.get('name')
                field_label = field.get('label', field_name)
                if field_name in case and case[field_name]:
                    value = str(case[field_name])[:100]
                    description_info.append(f"{field_label}: {value}...")
        
        print("📋 Description Information Found:")
        if description_info:
            for info in description_info:
                print(f"   {info}")
        else:
            print("   ❌ No description content found in any fields")
        
        # Show all field values for debugging
        print("🔍 All Description-Related Field Values:")
        for field_name in fields:
            if 'description' in field_name.lower() or 'brief' in field_name.lower() or 'detail' in field_name.lower() or 'summary' in field_name.lower():
                value = case.get(field_name)
                print(f"   {field_name}: {repr(value)}")
        
        print(f"✅ Found case: {cases[0]['CaseNumber']} - {cases[0]['Subject']}")
        
        # Get all related data for this case
        case_ids = [case['Id'] for case in cases]
        
        # Get complete Chatter feed data
        print("🔍 Retrieving Chatter feed...")
        feed_items = self._get_case_feed(case_ids)
        
        print("🔍 Retrieving feed comments...")
        feed_comments = self._get_feed_comments(case_ids)
        
        print("🔍 Retrieving feed tracked changes...")
        feed_tracked_changes = self._get_feed_tracked_changes(case_ids)
        
        print("🔍 Retrieving content documents...")
        content_documents = self._get_case_content_documents(case_ids)
        
        # Get all case-related objects
        print("🔍 Retrieving case articles...")
        articles = self._get_case_articles(case_ids)
        
        print("🔍 Retrieving case contact roles...")
        contact_roles = self._get_case_contact_roles(case_ids)
        
        print("🔍 Retrieving case milestones...")
        milestones = self._get_case_milestones(case_ids)
        
        print("🔍 Retrieving case participants...")
        participants = self._get_case_participants(case_ids)
        
        print("🔍 Retrieving case shares...")
        shares = self._get_case_shares(case_ids)
        
        print("🔍 Retrieving case solutions...")
        solutions = self._get_case_solutions(case_ids)
        
        print("🔍 Retrieving case tags...")
        tags = self._get_case_tags(case_ids)
        
        if include_comments:
            print("🔍 Retrieving case comments...")
            comments = self._get_case_comments(case_ids)
            
        if include_history:
            print("🔍 Retrieving case history...")
            history = self._get_case_history(case_ids)
        
        # Attach all data to the case
        for case in cases:
            case_id = case['Id']
            
            # Get feed items for this case and enrich with comments and tracked changes
            case_feed_items = feed_items.get(case_id, [])
            
            # Enrich each feed item with its comments and tracked changes
            for feed_item in case_feed_items:
                feed_item_id = feed_item['Id']
                feed_item['Comments'] = feed_comments.get(feed_item_id, [])
                feed_item['TrackedChanges'] = feed_tracked_changes.get(feed_item_id, [])
            
            case['ChatterFeed'] = case_feed_items
            case['FeedItemCount'] = len(case_feed_items)
            case['ContentDocuments'] = content_documents.get(case_id, [])
            
            # Add all case-related objects
            case['Articles'] = articles.get(case_id, [])
            case['ContactRoles'] = contact_roles.get(case_id, [])
            case['Milestones'] = milestones.get(case_id, [])
            case['Participants'] = participants.get(case_id, [])
            case['Shares'] = shares.get(case_id, [])
            case['Solutions'] = solutions.get(case_id, [])
            case['Tags'] = tags.get(case_id, [])
            
            if include_comments:
                case['Comments'] = comments.get(case_id, [])
            
            if include_history:
                case['History'] = history.get(case_id, [])
        
        # Apply AI cleaning to subjects if enabled
        if self.enable_ai_cleaning and cases:
            print("🤖 Applying AI cleaning to case subject...")
            for case in cases:
                original_subject = case.get('Subject', '')
                if original_subject:
                    case['OriginalSubject'] = original_subject
                    case['Subject'] = self.clean_subject_with_ai(original_subject)
        
        print(f"✅ Retrieved case {case_number} with complete data")
        return cases
    
    def _execute_query(self, query: str) -> List[Dict]:
        """Execute a SOQL query and return results."""
        headers = {
            'Authorization': f'Bearer {self.session_id}',
            'Content-Type': 'application/json'
        }
        
        url = f"{self.instance_url}/services/data/v58.0/query"
        params = {'q': query}
        
        try:
            print(f"🔍 Executing query: {query}")
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            
            data = response.json()
            records = data['records']
            
            # Handle pagination if more records exist
            while not data['done'] and 'nextRecordsUrl' in data:
                next_url = f"{self.instance_url}{data['nextRecordsUrl']}"
                response = requests.get(next_url, headers=headers)
                response.raise_for_status()
                data = response.json()
                records.extend(data['records'])
            
            return records
            
        except requests.exceptions.RequestException as e:
            print(f"❌ Failed to execute query: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response: {e.response.text}")
            return []
    
    def _get_case_comments(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get case comments for the specified case IDs."""
        if not case_ids:
            return {}
        
        # Build query for comments
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, ParentId, CommentBody, CreatedById, CreatedBy.Name, 
               CreatedDate, LastModifiedDate, IsPublished
        FROM CaseComment 
        WHERE ParentId IN ('{case_ids_str}')
        ORDER BY ParentId, CreatedDate
        """
        
        comments = self._execute_query(query)
        
        # Group comments by case ID
        comments_by_case = {}
        for comment in comments:
            case_id = comment['ParentId']
            if case_id not in comments_by_case:
                comments_by_case[case_id] = []
            
            # Clean up the comment data
            clean_comment = {
                'Id': comment['Id'],
                'Body': comment['CommentBody'],
                'CreatedBy': comment['CreatedBy']['Name'] if comment.get('CreatedBy') else None,
                'CreatedDate': comment['CreatedDate'],
                'LastModifiedDate': comment['LastModifiedDate'],
                'IsPublished': comment.get('IsPublished', False)
            }
            comments_by_case[case_id].append(clean_comment)
        
        return comments_by_case
    
    def _get_case_history(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get case history (field changes) for the specified case IDs."""
        if not case_ids:
            return {}
        
        # Build query for case history
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CaseId, Field, OldValue, NewValue, 
               CreatedById, CreatedBy.Name, CreatedDate
        FROM CaseHistory 
        WHERE CaseId IN ('{case_ids_str}')
        ORDER BY CaseId, CreatedDate
        """
        
        history_records = self._execute_query(query)
        
        # Group history by case ID
        history_by_case = {}
        for record in history_records:
            case_id = record['CaseId']
            if case_id not in history_by_case:
                history_by_case[case_id] = []
            
            # Clean up the history data
            clean_record = {
                'Id': record['Id'],
                'Field': record['Field'],
                'OldValue': record['OldValue'],
                'NewValue': record['NewValue'],
                'CreatedBy': record['CreatedBy']['Name'] if record.get('CreatedBy') else None,
                'CreatedDate': record['CreatedDate']
            }
            history_by_case[case_id].append(clean_record)
        
        return history_by_case
    
    def _get_case_feed(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get complete Chatter feed for cases - all feed posts including text content."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        
        # Enhanced query to get all feed item details (using only valid fields)
        query = f"""
        SELECT Id, Type, Body, Title, CreatedBy.Name, CreatedBy.Id, CreatedDate, ParentId,
               LikeCount, CommentCount, IsDeleted, InsertedById, Visibility,
               RelatedRecordId, LinkUrl
        FROM CaseFeed
        WHERE ParentId IN ('{case_ids_str}')
        ORDER BY ParentId, CreatedDate DESC
        """
        
        feed_items = self._execute_query(query)
        print(f"🔍 Found {len(feed_items)} feed items in CaseFeed")
        
        # Also try to get from FeedItem directly (sometimes contains different data)
        feeditem_query = f"""
        SELECT Id, Type, Body, Title, CreatedBy.Name, CreatedBy.Id, CreatedDate, ParentId,
               LikeCount, CommentCount, IsDeleted, InsertedById, Visibility,
               RelatedRecordId, LinkUrl
        FROM FeedItem
        WHERE ParentId IN ('{case_ids_str}')
        ORDER BY ParentId, CreatedDate DESC
        """
        
        try:
            additional_feed_items = self._execute_query(feeditem_query)
            print(f"🔍 Found {len(additional_feed_items)} additional feed items in FeedItem")
            
            # Combine both results, avoiding duplicates
            existing_ids = {item['Id'] for item in feed_items}
            for item in additional_feed_items:
                if item['Id'] not in existing_ids:
                    feed_items.append(item)
            
            print(f"🔍 Total unique feed items: {len(feed_items)}")
        except Exception as e:
            print(f"⚠️  Could not query FeedItem: {e}")
        
        # Group by case ID
        feed_by_case = {}
        post_types = {}
        
        for item in feed_items:
            case_id = item['ParentId']
            if case_id not in feed_by_case:
                feed_by_case[case_id] = []
            
            # Track post types for debugging
            post_type = item['Type']
            post_types[post_type] = post_types.get(post_type, 0) + 1
            
            clean_item = {
                'Id': item['Id'],
                'Type': item['Type'],
                'Body': item.get('Body', ''),
                'Title': item.get('Title', ''),
                'CreatedBy': item['CreatedBy']['Name'] if item.get('CreatedBy') else None,
                'CreatedById': item['CreatedBy']['Id'] if item.get('CreatedBy') else None,
                'CreatedDate': item['CreatedDate'],
                'LikeCount': item.get('LikeCount', 0),
                'CommentCount': item.get('CommentCount', 0),
                'IsDeleted': item.get('IsDeleted', False),
                'Visibility': item.get('Visibility', ''),
                'RelatedRecordId': item.get('RelatedRecordId', ''),
                'LinkUrl': item.get('LinkUrl', '')
            }
            feed_by_case[case_id].append(clean_item)
        
        # Show post type breakdown
        print("📊 Post Types Found:")
        for post_type, count in sorted(post_types.items()):
            print(f"   {post_type}: {count}")
        
        return feed_by_case
    
    def _get_feed_comments(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get comments on feed items for cases."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CommentBody, CreatedBy.Name, CreatedDate, FeedItemId
        FROM FeedComment
        WHERE FeedItemId IN (
            SELECT Id FROM CaseFeed WHERE ParentId IN ('{case_ids_str}')
        )
        ORDER BY FeedItemId, CreatedDate
        """
        
        try:
            comments = self._execute_query(query)
            print(f"🔍 Found {len(comments)} feed comments")
            
            # Group by feed item ID
            comments_by_feed_item = {}
            for comment in comments:
                feed_item_id = comment['FeedItemId']
                if feed_item_id not in comments_by_feed_item:
                    comments_by_feed_item[feed_item_id] = []
                
                clean_comment = {
                    'Id': comment['Id'],
                    'CommentBody': comment['CommentBody'],
                    'CreatedBy': comment['CreatedBy']['Name'] if comment.get('CreatedBy') else None,
                    'CreatedDate': comment['CreatedDate']
                }
                comments_by_feed_item[feed_item_id].append(clean_comment)
            
            return comments_by_feed_item
        except Exception as e:
            print(f"⚠️  Could not query FeedComment: {e}")
            return {}
    
    def _get_feed_tracked_changes(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get tracked field changes shown in feed."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, FieldName, OldValue, NewValue, CreatedDate, FeedItemId
        FROM FeedTrackedChange
        WHERE FeedItemId IN (
            SELECT Id FROM CaseFeed WHERE ParentId IN ('{case_ids_str}')
        )
        ORDER BY FeedItemId, CreatedDate
        """
        
        try:
            changes = self._execute_query(query)
            print(f"🔍 Found {len(changes)} feed tracked changes")
            
            # Group by feed item ID
            changes_by_feed_item = {}
            for change in changes:
                feed_item_id = change['FeedItemId']
                if feed_item_id not in changes_by_feed_item:
                    changes_by_feed_item[feed_item_id] = []
                
                clean_change = {
                    'Id': change['Id'],
                    'FieldName': change['FieldName'],
                    'OldValue': change['OldValue'],
                    'NewValue': change['NewValue'],
                    'CreatedDate': change['CreatedDate']
                }
                changes_by_feed_item[feed_item_id].append(clean_change)
            
            return changes_by_feed_item
        except Exception as e:
            print(f"⚠️  Could not query FeedTrackedChange: {e}")
            return {}
    
    def _get_case_content_documents(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get files attached to cases and feed posts."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, Title, FileType, ContentSize, CreatedBy.Name, CreatedDate
        FROM ContentDocument
        WHERE Id IN (
            SELECT ContentDocumentId
            FROM ContentDocumentLink
            WHERE LinkedEntityId IN ('{case_ids_str}')
               OR LinkedEntityId IN (
                  SELECT Id FROM CaseFeed WHERE ParentId IN ('{case_ids_str}')
               )
        )
        ORDER BY CreatedDate DESC
        """
        
        try:
            documents = self._execute_query(query)
            print(f"🔍 Found {len(documents)} content documents")
            
            # For simplicity, return all documents for each case
            # In practice, you might want to link them to specific feed items
            documents_by_case = {}
            for case_id in case_ids:
                documents_by_case[case_id] = []
                
            for doc in documents:
                clean_doc = {
                    'Id': doc['Id'],
                    'Title': doc['Title'],
                    'FileType': doc['FileType'],
                    'ContentSize': doc.get('ContentSize', 0),
                    'CreatedBy': doc['CreatedBy']['Name'] if doc.get('CreatedBy') else None,
                    'CreatedDate': doc['CreatedDate']
                }
                # Add to all cases (could be more specific with additional queries)
                for case_id in case_ids:
                    documents_by_case[case_id].append(clean_doc)
            
            return documents_by_case
        except Exception as e:
            print(f"⚠️  Could not query ContentDocument: {e}")
            return {}
    
    def _get_case_articles(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get knowledge articles associated with cases."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CaseId, KnowledgeArticleId, KnowledgeArticleVersionId, 
               CreatedById, CreatedBy.Name, CreatedDate, IsDeleted
        FROM CaseArticle 
        WHERE CaseId IN ('{case_ids_str}')
        AND IsDeleted = false
        ORDER BY CaseId, CreatedDate
        """
        
        try:
            articles = self._execute_query(query)
            print(f"🔍 Found {len(articles)} case articles")
            
            articles_by_case = {}
            for article in articles:
                case_id = article['CaseId']
                if case_id not in articles_by_case:
                    articles_by_case[case_id] = []
                
                clean_article = {
                    'Id': article['Id'],
                    'KnowledgeArticleId': article['KnowledgeArticleId'],
                    'KnowledgeArticleVersionId': article['KnowledgeArticleVersionId'],
                    'CreatedBy': article['CreatedBy']['Name'] if article.get('CreatedBy') else None,
                    'CreatedDate': article['CreatedDate']
                }
                articles_by_case[case_id].append(clean_article)
            
            return articles_by_case
        except Exception as e:
            print(f"⚠️  Could not query CaseArticle: {e}")
            return {}
    
    def _get_case_contact_roles(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get contact roles associated with cases."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CaseId, ContactId, Contact.Name, Role, 
               CreatedById, CreatedBy.Name, CreatedDate
        FROM CaseContactRole 
        WHERE CaseId IN ('{case_ids_str}')
        ORDER BY CaseId, CreatedDate
        """
        
        try:
            roles = self._execute_query(query)
            print(f"🔍 Found {len(roles)} case contact roles")
            
            roles_by_case = {}
            for role in roles:
                case_id = role['CaseId']
                if case_id not in roles_by_case:
                    roles_by_case[case_id] = []
                
                clean_role = {
                    'Id': role['Id'],
                    'ContactId': role['ContactId'],
                    'ContactName': role['Contact']['Name'] if role.get('Contact') else None,
                    'Role': role['Role'],
                    'CreatedBy': role['CreatedBy']['Name'] if role.get('CreatedBy') else None,
                    'CreatedDate': role['CreatedDate']
                }
                roles_by_case[case_id].append(clean_role)
            
            return roles_by_case
        except Exception as e:
            print(f"⚠️  Could not query CaseContactRole: {e}")
            return {}
    
    def _get_case_milestones(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get case milestones."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CaseId, MilestoneTypeId, MilestoneType.Name, 
               StartDate, TargetDate, CompletionDate, IsCompleted, IsViolated,
               CreatedDate
        FROM CaseMilestone 
        WHERE CaseId IN ('{case_ids_str}')
        ORDER BY CaseId, CreatedDate
        """
        
        try:
            milestones = self._execute_query(query)
            print(f"🔍 Found {len(milestones)} case milestones")
            
            milestones_by_case = {}
            for milestone in milestones:
                case_id = milestone['CaseId']
                if case_id not in milestones_by_case:
                    milestones_by_case[case_id] = []
                
                clean_milestone = {
                    'Id': milestone['Id'],
                    'MilestoneTypeId': milestone['MilestoneTypeId'],
                    'MilestoneTypeName': milestone['MilestoneType']['Name'] if milestone.get('MilestoneType') else None,
                    'StartDate': milestone['StartDate'],
                    'TargetDate': milestone['TargetDate'],
                    'CompletionDate': milestone['CompletionDate'],
                    'IsCompleted': milestone['IsCompleted'],
                    'IsViolated': milestone['IsViolated'],
                    'CreatedDate': milestone['CreatedDate']
                }
                milestones_by_case[case_id].append(clean_milestone)
            
            return milestones_by_case
        except Exception as e:
            print(f"⚠️  Could not query CaseMilestone: {e}")
            return {}
    
    def _get_case_participants(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get case participants."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CaseId, UserId, User.Name, Role, 
               CreatedById, CreatedBy.Name, CreatedDate
        FROM CaseParticipant 
        WHERE CaseId IN ('{case_ids_str}')
        ORDER BY CaseId, CreatedDate
        """
        
        try:
            participants = self._execute_query(query)
            print(f"🔍 Found {len(participants)} case participants")
            
            participants_by_case = {}
            for participant in participants:
                case_id = participant['CaseId']
                if case_id not in participants_by_case:
                    participants_by_case[case_id] = []
                
                clean_participant = {
                    'Id': participant['Id'],
                    'UserId': participant['UserId'],
                    'UserName': participant['User']['Name'] if participant.get('User') else None,
                    'Role': participant['Role'],
                    'CreatedBy': participant['CreatedBy']['Name'] if participant.get('CreatedBy') else None,
                    'CreatedDate': participant['CreatedDate']
                }
                participants_by_case[case_id].append(clean_participant)
            
            return participants_by_case
        except Exception as e:
            print(f"⚠️  Could not query CaseParticipant: {e}")
            return {}
    
    def _get_case_shares(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get case sharing records."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CaseId, UserOrGroupId, CaseAccessLevel, RowCause,
               LastModifiedById, LastModifiedBy.Name, LastModifiedDate
        FROM CaseShare 
        WHERE CaseId IN ('{case_ids_str}')
        ORDER BY CaseId, LastModifiedDate
        """
        
        try:
            shares = self._execute_query(query)
            print(f"🔍 Found {len(shares)} case shares")
            
            shares_by_case = {}
            for share in shares:
                case_id = share['CaseId']
                if case_id not in shares_by_case:
                    shares_by_case[case_id] = []
                
                clean_share = {
                    'Id': share['Id'],
                    'UserOrGroupId': share['UserOrGroupId'],
                    'CaseAccessLevel': share['CaseAccessLevel'],
                    'RowCause': share['RowCause'],
                    'LastModifiedBy': share['LastModifiedBy']['Name'] if share.get('LastModifiedBy') else None,
                    'LastModifiedDate': share['LastModifiedDate']
                }
                shares_by_case[case_id].append(clean_share)
            
            return shares_by_case
        except Exception as e:
            print(f"⚠️  Could not query CaseShare: {e}")
            return {}
    
    def _get_case_solutions(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get case solutions."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, CaseId, SolutionId, Solution.SolutionName, Solution.SolutionNote,
               CreatedById, CreatedBy.Name, CreatedDate
        FROM CaseSolution 
        WHERE CaseId IN ('{case_ids_str}')
        ORDER BY CaseId, CreatedDate
        """
        
        try:
            solutions = self._execute_query(query)
            print(f"🔍 Found {len(solutions)} case solutions")
            
            solutions_by_case = {}
            for solution in solutions:
                case_id = solution['CaseId']
                if case_id not in solutions_by_case:
                    solutions_by_case[case_id] = []
                
                clean_solution = {
                    'Id': solution['Id'],
                    'SolutionId': solution['SolutionId'],
                    'SolutionName': solution['Solution']['SolutionName'] if solution.get('Solution') else None,
                    'SolutionNote': solution['Solution']['SolutionNote'] if solution.get('Solution') else None,
                    'CreatedBy': solution['CreatedBy']['Name'] if solution.get('CreatedBy') else None,
                    'CreatedDate': solution['CreatedDate']
                }
                solutions_by_case[case_id].append(clean_solution)
            
            return solutions_by_case
        except Exception as e:
            print(f"⚠️  Could not query CaseSolution: {e}")
            return {}
    
    def _get_case_tags(self, case_ids: List[str]) -> Dict[str, List[Dict]]:
        """Get case tags."""
        if not case_ids:
            return {}
        
        case_ids_str = "', '".join(case_ids)
        query = f"""
        SELECT Id, ItemId, TagDefinitionId, TagDefinition.Name,
               CreatedById, CreatedBy.Name, CreatedDate
        FROM CaseTag 
        WHERE ItemId IN ('{case_ids_str}')
        ORDER BY ItemId, CreatedDate
        """
        
        try:
            tags = self._execute_query(query)
            print(f"🔍 Found {len(tags)} case tags")
            
            tags_by_case = {}
            for tag in tags:
                case_id = tag['ItemId']
                if case_id not in tags_by_case:
                    tags_by_case[case_id] = []
                
                clean_tag = {
                    'Id': tag['Id'],
                    'TagDefinitionId': tag['TagDefinitionId'],
                    'TagName': tag['TagDefinition']['Name'] if tag.get('TagDefinition') else None,
                    'CreatedBy': tag['CreatedBy']['Name'] if tag.get('CreatedBy') else None,
                    'CreatedDate': tag['CreatedDate']
                }
                tags_by_case[case_id].append(clean_tag)
            
            return tags_by_case
        except Exception as e:
            print(f"⚠️  Could not query CaseTag: {e}")
            return {}
    

    
    def save_to_json(self, cases: List[Dict], filename: str = None) -> str:
        """
        Save cases to JSON file.
        
        Args:
            cases: List of case dictionaries
            filename: Output filename. If None, generates timestamp-based name.
            
        Returns:
            str: Filename of saved file
        """
        # Create output directory
        output_dir = "out"
        os.makedirs(output_dir, exist_ok=True)
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"salesforce_cases_{timestamp}.json"
        
        # Add output directory to filename
        filepath = os.path.join(output_dir, filename)
        
        # Remove Salesforce metadata from records
        clean_cases = []
        for case in cases:
            clean_case = {k: v for k, v in case.items() if k != 'attributes'}
            clean_cases.append(clean_case)
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(clean_cases, f, indent=2, default=str)
            
            print(f"✅ Saved {len(cases)} cases to {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Failed to save JSON file: {e}")
            return ""
    
    def save_to_csv(self, cases: List[Dict], filename: str = None) -> str:
        """
        Save cases to CSV file.
        
        Args:
            cases: List of case dictionaries
            filename: Output filename. If None, generates timestamp-based name.
            
        Returns:
            str: Filename of saved file
        """
        if not cases:
            print("❌ No cases to save")
            return ""
        
        # Create output directory
        output_dir = "out"
        os.makedirs(output_dir, exist_ok=True)
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"salesforce_cases_{timestamp}.csv"
        
        # Add output directory to filename
        filepath = os.path.join(output_dir, filename)
        
        try:
            # Get all unique field names
            fieldnames = set()
            for case in cases:
                fieldnames.update(k for k in case.keys() if k != 'attributes')
            fieldnames = sorted(list(fieldnames))
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for case in cases:
                    # Remove Salesforce metadata and write row
                    clean_case = {k: v for k, v in case.items() if k != 'attributes'}
                    writer.writerow(clean_case)
            
            print(f"✅ Saved {len(cases)} cases to {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Failed to save CSV file: {e}")
            return ""
    
    def save_to_excel(self, cases: List[Dict], filename: str = None) -> str:
        """
        Save cases to Excel file with hyperlinked case numbers.
        
        Args:
            cases: List of case dictionaries
            filename: Output filename. If None, generates timestamp-based name.
            
        Returns:
            str: Filename of saved file
        """
        if not cases:
            print("❌ No cases to save")
            return ""
        
        # Create output directory
        output_dir = "out"
        os.makedirs(output_dir, exist_ok=True)
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"salesforce_cases_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename = filename.replace('.json', '.xlsx').replace('.csv', '.xlsx')
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
        
        # Add output directory to filename
        filepath = os.path.join(output_dir, filename)
        
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Salesforce Cases"
            
            # Set headers
            headers = ['CaseNumber', 'Subject']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            for row_idx, case in enumerate(cases, 2):
                case_number = case.get('CaseNumber', '')
                case_id = case.get('Id', '')
                subject = case.get('Subject', '')
                
                # Case Number with hyperlink
                case_cell = ws.cell(row=row_idx, column=1, value=case_number)
                if case_id:
                    # Create Salesforce URL
                    salesforce_url = f"https://vastdata.lightning.force.com/lightning/r/Case/{case_id}/view"
                    case_cell.hyperlink = salesforce_url
                    case_cell.font = Font(color="0000FF", underline="single")  # Blue underlined link
                
                # Subject
                subject_cell = ws.cell(row=row_idx, column=2, value=subject)
                
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                
                # Set column width with some padding
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            
            print(f"✅ Saved {len(cases)} cases to {filepath}")
            print(f"📊 Excel file contains:")
            print(f"   - CaseNumber column with clickable links to Salesforce")
            print(f"   - Subject column with cleaned subjects (if AI cleaning was enabled)")
            return filepath
            
        except Exception as e:
            print(f"❌ Failed to save Excel file: {e}")
            return ""

def main():
    """
    Salesforce Cases Client with command line interface
    """
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Retrieve Salesforce case data')
    parser.add_argument('--cid', '--case-id', dest='case_id', 
                       help='Case number to retrieve (e.g., 00081000)')
    parser.add_argument('--all', action='store_true',
                       help='Retrieve all cases (All Cases - Vast Data list)')
    parser.add_argument('--limit', type=int, default=100,
                       help='Maximum number of cases to retrieve (default: 100)')
    parser.add_argument('--status', 
                       help='Filter by case status (e.g., New, Working, Solved)')
    parser.add_argument('--exclude-merged', action='store_true',
                       help='Exclude cases with "Merged" in the subject (status is solved AND not Merged)')
    parser.add_argument('--list-only', action='store_true',
                       help='Get only basic case list (no detailed data, feed, or custom fields)')
    parser.add_argument('--comments', action='store_true', default=True,
                       help='Include case comments (default: True)')
    parser.add_argument('--history', action='store_true', default=False,
                       help='Include case history (default: False)')
    parser.add_argument('--clean-subjects', action='store_true', default=False,
                       help='Enable AI-powered subject cleaning using OpenAI (requires OpenAI API key)')
    parser.add_argument('--excel', action='store_true', default=False,
                       help='Generate Excel file (.xlsx) with hyperlinked case numbers')
    
    args = parser.parse_args()
    
    # Configuration - Replace with your actual credentials
    USERNAME = "yong.li@vastdata.com"
    PASSWORD = "fav8@Apple!_v1"
    SECURITY_TOKEN = "TRPXpOSnfjA2WCrfDiewUIns"  # Add your actual security token here
    IS_SANDBOX = False  # Set to True if using sandbox
    
    # Check if security token is provided
    if not SECURITY_TOKEN:
        print("❌ Please set your SECURITY_TOKEN in the script")
        print("   You can find your security token in Salesforce:")
        print("   1. Go to Setup > My Personal Information > Reset My Security Token")
        print("   2. Check your email for the new token")
        print("   3. Update the SECURITY_TOKEN variable in this script")
        return
    
    # Initialize client
    client = SalesforceCasesClient(
        username=USERNAME,
        password=PASSWORD,
        security_token=SECURITY_TOKEN,
        is_sandbox=IS_SANDBOX,
        enable_ai_cleaning=args.clean_subjects
    )
    
    # Authenticate
    if not client.authenticate():
        print("❌ Authentication failed. Please check your credentials and try again.")
        return
    
    # Determine what to retrieve
    if args.all:
        # Build filter description
        filter_desc = f"limit: {args.limit}"
        if args.status:
            filter_desc += f", status: {args.status}"
        if args.exclude_merged:
            filter_desc += ", excluding merged cases"
        if args.list_only:
            filter_desc += ", list only"
            
        print(f"🔍 Retrieving all cases from Salesforce ({filter_desc})...")
        
        if args.list_only:
            # Get basic case list only (fast)
            cases = client.get_cases_list_only(
                limit=args.limit, 
                status_filter=args.status,
                exclude_merged=args.exclude_merged
            )
        else:
            # Get full case data with custom fields and feed
            print("🔍 Discovering custom description fields...")
            description_fields = client.describe_case_object()
            
            cases = client.get_cases(
                limit=args.limit, 
                status_filter=args.status,
                exclude_merged=args.exclude_merged,
                include_comments=args.comments, 
                include_history=args.history
            )
    else:
        # Get specific case by number
        case_number = args.case_id
        if not case_number:
            case_number = input("Enter the case number to retrieve (e.g., 00078683): ").strip()
        
        if not case_number:
            print("❌ No case number provided. Exiting.")
            print("Usage: python3 sfc.py --cid 00081000 or python3 sfc.py --all")
            return
        
        # Check if case file already exists (skip if it does)
        clean_case_number = case_number.lstrip('0') or '0'
        output_file = f"out/{clean_case_number}.json"
        
        if os.path.exists(output_file):
            print(f"✅ Case file already exists: {output_file}")
            print(f"🔍 Skipping fetch for case {case_number} (already downloaded)")
            print(f"💡 To re-fetch, delete the file: rm {output_file}")
            return
        
        if args.list_only:
            print(f"🔍 Retrieving case {case_number} (basic info only)...")
            # For single case, convert to list format
            single_case = client.get_cases_list_only(limit=1, status_filter=None, exclude_merged=False)
            cases = [case for case in single_case if case.get('CaseNumber') == case_number]
        else:
            print("🔍 Discovering custom description fields...")
            description_fields = client.describe_case_object()
            
            print(f"🔍 Retrieving case {case_number} from Salesforce...")
            cases = client.get_cases_by_number(case_number, include_comments=args.comments, include_history=args.history, custom_description_fields=description_fields)
    
    if cases:
        # Determine filename based on query type
        if len(cases) == 1 and not args.all:
            # Single case: use case number as filename (remove leading zeros)
            case_number = cases[0].get('CaseNumber', '')
            clean_case_number = case_number.lstrip('0') or '0'  # Keep at least one zero if all zeros
            filename = f"{clean_case_number}.json"
        elif args.all and args.status == "Solved" and args.exclude_merged:
            # Solved cases list: use solved.json
            filename = "solved.json"
        else:
            # Other lists: use default timestamp-based name
            filename = None
        
        # Save files based on options
        json_file = client.save_to_json(cases, filename)
        
        files_created = [f"JSON: {json_file}"]
        
        # Generate Excel file if requested
        if args.excel:
            excel_filename = filename.replace('.json', '.xlsx') if filename else None
            excel_file = client.save_to_excel(cases, excel_filename)
            if excel_file:
                files_created.append(f"Excel: {excel_file}")
        
        print(f"\n📄 Files created:")
        for file_info in files_created:
            print(f"   {file_info}")
        
        # Display summary
        print(f"\n📊 Summary:")
        print(f"   Total cases: {len(cases)}")
        
        # Count by status
        status_counts = {}
        for case in cases:
            status = case.get('Status', 'Unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        print(f"   By Status:")
        for status, count in sorted(status_counts.items()):
            print(f"     {status}: {count}")
            
        # Show filename format for single cases
        if len(cases) == 1 and not args.all:
            print(f"   Case file: {clean_case_number}.json (leading zeros removed)")
    else:
        print("❌ No cases retrieved. Please check your Salesforce permissions and query.")

if __name__ == "__main__":
    main()